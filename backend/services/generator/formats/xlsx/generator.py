"""
XLSX Format Generator

Full implementation of Excel spreadsheet generation.
Migrated from generator.py for modularity.
"""

import os
import re
from typing import Optional, TYPE_CHECKING

import structlog

from ..base import BaseFormatGenerator
from ..factory import register_generator
from ...models import OutputFormat, GenerationJob
from ...config import THEMES

if TYPE_CHECKING:
    from ...template_analyzer import TemplateAnalysis

logger = structlog.get_logger(__name__)


# Sheet review constraints (Excel limitations)
XLSX_CONSTRAINTS = {
    "sheet_name_max_chars": 31,  # Excel hard limit
    "cell_max_chars": 32767,  # Excel cell limit
    "header_max_chars": 50,  # Recommended for readability
    "content_cell_max_chars": 500,  # Recommended for readability
}


def get_theme_colors(theme_key: str = "business", custom_colors: dict = None) -> dict:
    """Get theme colors, with fallback to business theme."""
    theme = THEMES.get(theme_key, THEMES["business"]).copy()
    if custom_colors:
        for key in ["primary", "secondary", "accent", "text", "background"]:
            if key in custom_colors:
                theme[key] = custom_colors[key]
    return theme


def strip_markdown(text: str) -> str:
    """Remove markdown formatting, returning clean text."""
    if not text:
        return ""
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    text = re.sub(r'__([^_]+)__', r'\1', text)
    text = re.sub(r'_([^_]+)_', r'\1', text)
    text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)
    text = re.sub(r'```[^`]*```', '', text, flags=re.DOTALL)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    return text


@register_generator(OutputFormat.XLSX)
class XLSXGenerator(BaseFormatGenerator):
    """Full implementation of Excel spreadsheet generator.

    Features:
    - Professional styling with themes
    - Multiple sheets (Summary, Content, Sources)
    - Cell formatting with colors
    - Auto column width adjustment
    """

    @property
    def format_name(self) -> str:
        return "xlsx"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    async def generate(
        self,
        job: GenerationJob,
        filename: str,
        template_analysis: Optional["TemplateAnalysis"] = None,
    ) -> str:
        """Generate an Excel spreadsheet.

        Args:
            job: The generation job containing metadata and sections
            filename: The output filename
            template_analysis: Optional template analysis for styling

        Returns:
            Path to the generated XLSX file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # Get output directory from config
            from ...models import GenerationConfig
            config = GenerationConfig()

            # Determine include_sources
            include_sources = job.metadata.get("include_sources", config.include_sources)

            wb = Workbook()

            # Get theme colors
            theme_key = job.metadata.get("theme", "business")
            custom_colors = job.metadata.get("custom_colors")
            theme = get_theme_colors(theme_key, custom_colors)

            # Convert theme color (without #)
            primary_hex = theme["primary"].lstrip('#')

            # Font mapping
            XLSX_FONT_MAP = {
                "modern": "Calibri",
                "classic": "Times New Roman",
                "professional": "Arial",
                "technical": "Consolas",
            }
            font_family_key = job.metadata.get("font_family", "modern")
            excel_font_name = XLSX_FONT_MAP.get(font_family_key, XLSX_FONT_MAP["modern"])

            # Define styles
            header_font = Font(name=excel_font_name, bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color=primary_hex, end_color=primary_hex, fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # ========== Sheet 1: Summary ==========
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Title row
            ws_summary.merge_cells('A1:B1')
            ws_summary['A1'] = job.title
            ws_summary['A1'].font = Font(name=excel_font_name, bold=True, size=16)
            ws_summary['A1'].alignment = Alignment(horizontal="center")

            # Summary data
            summary_data = [
                ("Description", job.outline.description if job.outline else ""),
                ("Status", job.status.value),
                ("Created", job.created_at.strftime("%Y-%m-%d %H:%M")),
                ("Completed", job.completed_at.strftime("%Y-%m-%d %H:%M") if job.completed_at else "In Progress"),
                ("Total Sections", str(len(job.sections))),
                ("Sources Used", str(len(job.sources_used))),
            ]

            for i, (label, value) in enumerate(summary_data, start=3):
                ws_summary[f'A{i}'] = label
                ws_summary[f'A{i}'].font = Font(name=excel_font_name, bold=True)
                ws_summary[f'B{i}'] = value
                ws_summary[f'A{i}'].border = thin_border
                ws_summary[f'B{i}'].border = thin_border

            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 60

            # ========== Sheet Review Setup ==========
            # Optional LLM-based quality review for content
            enable_quality_review = job.metadata.get("enable_slide_review", False) or job.metadata.get("enable_quality_review", False)
            sheet_reviewer = None
            if enable_quality_review:
                try:
                    from .reviewer import XLSXSheetReviewer
                    llm_generate_func = job.metadata.get("llm_generate_func")
                    sheet_reviewer = XLSXSheetReviewer(llm_generate_func=llm_generate_func)
                    logger.info("XLSX sheet reviewer enabled")
                except ImportError:
                    logger.warning("XLSXSheetReviewer not available, skipping quality review")

            # ========== Sheet 2: Content ==========
            ws_content = wb.create_sheet("Content")

            # Headers
            headers = ["Section #", "Title", "Content", "Approved", "Feedback"]
            for col, header in enumerate(headers, start=1):
                cell = ws_content.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Review content sheet structure if enabled
            if sheet_reviewer:
                content_sheet_spec = {
                    "name": "Content",
                    "headers": headers,
                    "rows": [
                        {
                            "cells": [
                                section.order,
                                section.title,
                                strip_markdown(section.revised_content or section.content or ""),
                                "Yes" if section.approved else "No",
                                section.feedback or ""
                            ]
                        }
                        for section in job.sections
                    ]
                }
                try:
                    review_result, fix_result = await sheet_reviewer.review_and_fix(
                        content_sheet_spec, 0, XLSX_CONSTRAINTS
                    )
                    if review_result.has_issues:
                        logger.warning(
                            "Content sheet has issues",
                            issues=[issue.description for issue in review_result.issues]
                        )
                    if fix_result.fixes_applied > 0:
                        logger.info(f"Applied {fix_result.fixes_applied} fixes to content sheet")
                except Exception as e:
                    logger.warning(f"Sheet review failed: {e}")

            # Content rows
            for i, section in enumerate(job.sections, start=2):
                content = section.revised_content or section.content
                content = strip_markdown(content) if content else ""

                ws_content.cell(row=i, column=1, value=section.order).border = thin_border
                ws_content.cell(row=i, column=2, value=section.title).border = thin_border
                ws_content.cell(row=i, column=3, value=content).border = thin_border
                ws_content.cell(row=i, column=3).alignment = cell_alignment
                ws_content.cell(row=i, column=4, value="Yes" if section.approved else "No").border = thin_border
                ws_content.cell(row=i, column=5, value=section.feedback or "").border = thin_border

            ws_content.column_dimensions['A'].width = 12
            ws_content.column_dimensions['B'].width = 30
            ws_content.column_dimensions['C'].width = 80
            ws_content.column_dimensions['D'].width = 12
            ws_content.column_dimensions['E'].width = 40

            # ========== Sheet 3: Sources ==========
            if include_sources and job.sources_used:
                ws_sources = wb.create_sheet("Sources")

                # Headers - added Usage Type and Description columns
                source_headers = ["#", "Document Name", "Location", "Usage Type", "Description", "Relevance", "Snippet"]
                for col, header in enumerate(source_headers, start=1):
                    cell = ws_sources.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # Group sources by usage type for better organization
                from ...models import SourceUsageType
                content_sources = [s for s in job.sources_used if getattr(s, 'usage_type', SourceUsageType.CONTENT) == SourceUsageType.CONTENT]
                style_sources = [s for s in job.sources_used if getattr(s, 'usage_type', None) == SourceUsageType.STYLE]
                other_sources = [s for s in job.sources_used if getattr(s, 'usage_type', None) not in [SourceUsageType.CONTENT, SourceUsageType.STYLE, None]]

                # Combine in order: content first, then style, then other
                ordered_sources = content_sources + style_sources + other_sources
                if not ordered_sources:
                    ordered_sources = job.sources_used

                # Source rows
                for i, source in enumerate(ordered_sources, start=2):
                    ws_sources.cell(row=i, column=1, value=i-1).border = thin_border

                    # Document name cell with hyperlink if available
                    doc_name_cell = ws_sources.cell(row=i, column=2, value=source.document_name)
                    doc_name_cell.border = thin_border

                    # Add hyperlink if available
                    hyperlink_url = None
                    if hasattr(source, 'document_url') and source.document_url:
                        hyperlink_url = source.document_url
                    elif hasattr(source, 'document_path') and source.document_path:
                        import urllib.parse
                        hyperlink_url = f"file://{urllib.parse.quote(source.document_path)}"

                    if hyperlink_url:
                        try:
                            doc_name_cell.hyperlink = hyperlink_url
                            doc_name_cell.font = Font(name=excel_font_name, color="0000FF", underline="single")
                        except Exception:
                            pass  # Fall back to plain text

                    doc_name = source.document_name or ""
                    if source.page_number:
                        location_label = f"Slide {source.page_number}" if doc_name.lower().endswith('.pptx') else f"Page {source.page_number}"
                    else:
                        location_label = "N/A"
                    ws_sources.cell(row=i, column=3, value=location_label).border = thin_border

                    # Usage type
                    usage_type = getattr(source, 'usage_type', SourceUsageType.CONTENT)
                    usage_type_label = usage_type.value.title() if hasattr(usage_type, 'value') else str(usage_type)
                    ws_sources.cell(row=i, column=4, value=usage_type_label).border = thin_border

                    # Usage description
                    usage_desc = getattr(source, 'usage_description', '') or ''
                    ws_sources.cell(row=i, column=5, value=usage_desc).border = thin_border

                    ws_sources.cell(row=i, column=6, value=f"{source.relevance_score:.2f}").border = thin_border
                    ws_sources.cell(row=i, column=7, value=source.snippet[:200] + "..." if len(source.snippet) > 200 else source.snippet).border = thin_border
                    ws_sources.cell(row=i, column=7).alignment = cell_alignment

                ws_sources.column_dimensions['A'].width = 8
                ws_sources.column_dimensions['B'].width = 40
                ws_sources.column_dimensions['C'].width = 12
                ws_sources.column_dimensions['D'].width = 12
                ws_sources.column_dimensions['E'].width = 30
                ws_sources.column_dimensions['F'].width = 10
                ws_sources.column_dimensions['G'].width = 60

            # Save workbook
            output_path = os.path.join(config.output_dir, f"{filename}.xlsx")
            wb.save(output_path)

            logger.info("XLSX generated", path=output_path)
            return output_path

        except ImportError as e:
            logger.error(f"openpyxl import error: {e}")
            raise
